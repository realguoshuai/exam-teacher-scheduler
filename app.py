"""
Web Application for Exam Teacher Scheduler
"""

from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
from models import Teacher, Exam
from scheduler import ExamScheduler
from utils import init_data_dir, load_teachers, load_exams, export_schedule, save_teachers, save_exams, load_config, save_config
import config
import os
import sys
import pandas as pd
from datetime import datetime
import threading

# 处理 PyInstaller 打包后的路径问题
if getattr(sys, 'frozen', False):
    # 如果是打包后的exe
    application_path = os.path.dirname(sys.executable)
    template_folder = os.path.join(application_path, 'templates')
    app = Flask(__name__, template_folder=template_folder)
else:
    # 如果是开发环境
    app = Flask(__name__)
CORS(app)

app.config['SECRET_KEY'] = 'exam-teacher-scheduler-secret-key-2024'
app.config['JSON_AS_ASCII'] = False

scheduler_instance = None
scheduler_lock = threading.Lock()


def get_scheduler():
    """Get or create scheduler instance"""
    global scheduler_instance
    if scheduler_instance is None:
        teachers = load_teachers()
        exams = load_exams()
        config = load_config()
        if teachers and exams:
            scheduler_instance = ExamScheduler(teachers, exams, config)
    return scheduler_instance


def reset_scheduler():
    """Reset scheduler instance"""
    global scheduler_instance
    scheduler_instance = None


@app.route('/')
def index():
    """Home page"""
    return render_template('index.html')


@app.route('/api/health')
def health():
    """Health check"""
    return jsonify({'status': 'ok', 'timestamp': datetime.now().isoformat()})


@app.route('/api/init')
def init_data():
    """Initialize data"""
    try:
        init_data_dir()
        reset_scheduler()
        return jsonify({'success': True, 'message': 'Data initialized successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/config')
def api_config():
    """Get configuration"""
    try:
        config = load_config()
        return jsonify({'success': True, 'data': config})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/config', methods=['POST'])
def api_update_config():
    """Update configuration"""
    try:
        data = request.json
        config = load_config()
        for key, value in data.items():
            config[key] = value
        save_config(config)
        reset_scheduler()
        return jsonify({'success': True, 'message': 'Configuration updated successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/teachers')
def api_teachers():
    """Get all teachers"""
    try:
        teachers = load_teachers()
        
        # 从 final_schedules 计算每个老师的监考次数
        scheduler = get_scheduler()
        teacher_exam_count = {}
        if scheduler and scheduler.final_schedules:
            for schedule in scheduler.final_schedules:
                for teacher in schedule.teachers:
                    if teacher.teacher_id not in teacher_exam_count:
                        teacher_exam_count[teacher.teacher_id] = 0
                    teacher_exam_count[teacher.teacher_id] += 1
        
        teacher_list = []
        for t in teachers:
            exam_count = teacher_exam_count.get(t.teacher_id, 0)
            teacher_list.append({
                'id': t.teacher_id,
                'name': t.name,
                'title': t.title,
                'phone': t.phone,
                'department': t.department,
                'exam_count': exam_count
            })
        return jsonify({'success': True, 'data': teacher_list, 'count': len(teacher_list)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/teachers/import', methods=['POST'])
def api_import_teachers():
    """Import teachers from Excel file"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400

        if not file.filename.endswith('.xlsx'):
            return jsonify({'success': False, 'error': 'Only .xlsx files are supported'}), 400

        df = pd.read_excel(file)

        required_columns = ['工号', '姓名', '职称', '联系方式', '所属部门']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({'success': False, 'error': f'Missing columns: {missing_columns}'}), 400

        teachers = load_teachers()
        existing_ids = {t.teacher_id for t in teachers}
        added_count = 0

        for _, row in df.iterrows():
            teacher_id = str(row['工号'])
            if teacher_id not in existing_ids:
                teacher = Teacher(
                    teacher_id=teacher_id,
                    name=str(row['姓名']),
                    title=str(row['职称']),
                    phone=str(row['联系方式']),
                    department=str(row['所属部门']),
                    exam_count=0
                )
                teachers.append(teacher)
                added_count += 1

        save_teachers(teachers)
        reset_scheduler()

        return jsonify({
            'success': True,
            'message': f'Successfully imported {added_count} teachers',
            'count': added_count
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/teachers', methods=['POST'])
def api_add_teacher():
    """Add a new teacher manually"""
    try:
        data = request.json
        required_fields = ['id', 'name', 'title', 'phone', 'department']
        missing = [f for f in required_fields if f not in data]
        if missing:
            return jsonify({'success': False, 'error': f'Missing fields: {missing}'}), 400

        teachers = load_teachers()
        existing_ids = {t.teacher_id for t in teachers}
        if data['id'] in existing_ids:
            return jsonify({'success': False, 'error': 'Teacher ID already exists'}), 400

        teacher = Teacher(
            teacher_id=str(data['id']),
            name=data['name'],
            title=data['title'],
            phone=data['phone'],
            department=data['department'],
            exam_count=0
        )
        teachers.append(teacher)
        save_teachers(teachers)
        reset_scheduler()

        return jsonify({'success': True, 'message': 'Teacher added successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/teachers/<teacher_id>', methods=['DELETE'])
def api_delete_teacher(teacher_id):
    """Delete a teacher"""
    try:
        teachers = load_teachers()
        teachers = [t for t in teachers if t.teacher_id != teacher_id]
        save_teachers(teachers)
        reset_scheduler()
        return jsonify({'success': True, 'message': 'Teacher deleted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/exams')
def api_exams():
    """Get all exams"""
    try:
        exams = load_exams()
        exam_list = []
        for e in exams:
            exam_list.append({
                'id': e.exam_id,
                'name': e.exam_name,
                'subject': e.subject,
                'date': e.date,
                'time_slot': e.time_slot,
                'room': e.room,
                'required_teachers': e.required_teachers
            })
        return jsonify({'success': True, 'data': exam_list, 'count': len(exam_list)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/exams/import', methods=['POST'])
def api_import_exams():
    """Import exams from Excel file"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400

        if not file.filename.endswith('.xlsx'):
            return jsonify({'success': False, 'error': 'Only .xlsx files are supported'}), 400

        df = pd.read_excel(file)

        required_columns = ['考试编号', '考试名称', '科目', '日期', '时间段']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({'success': False, 'error': f'Missing columns: {missing_columns}'}), 400

        exams = load_exams()
        existing_ids = {e.exam_id for e in exams}
        added_count = 0

        for _, row in df.iterrows():
            exam_id = str(row['考试编号'])
            if exam_id not in existing_ids:
                rooms_count_value = row.get('考场数', 6)
                if rooms_count_value is None or str(rooms_count_value) != str(rooms_count_value):
                    rooms_count_value = 6

                exam = Exam(
                    exam_id=exam_id,
                    exam_name=str(row['考试名称']),
                    subject=str(row['科目']),
                    date=str(row['日期']),
                    time_slot=str(row['时间段']),
                    room=str(row.get('考场', '')),
                    required_teachers=int(row.get('需要监考人数', 2) or 2),
                    rooms_count=int(rooms_count_value)
                )
                exams.append(exam)
                added_count += 1

        save_exams(exams)
        reset_scheduler()

        return jsonify({
            'success': True,
            'message': f'Successfully imported {added_count} exams',
            'count': added_count
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/exams', methods=['POST'])
def api_add_exam():
    """Add a new exam manually"""
    try:
        data = request.json
        required_fields = ['id', 'name', 'subject', 'date', 'time_slot']
        missing = [f for f in required_fields if f not in data]
        if missing:
            return jsonify({'success': False, 'error': f'Missing fields: {missing}'}), 400

        exams = load_exams()
        existing_ids = {e.exam_id for e in exams}
        if data['id'] in existing_ids:
            return jsonify({'success': False, 'error': 'Exam ID already exists'}), 400

        exam = Exam(
            exam_id=str(data['id']),
            exam_name=data['name'],
            subject=data['subject'],
            date=data['date'],
            time_slot=data['time_slot'],
            room=data.get('room', ''),
            required_teachers=int(data.get('required_teachers', 2)),
            rooms_count=int(data.get('rooms_count', 6))
        )
        exams.append(exam)
        save_exams(exams)
        reset_scheduler()

        return jsonify({'success': True, 'message': 'Exam added successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/exams/<exam_id>', methods=['DELETE'])
def api_delete_exam(exam_id):
    """Delete an exam"""
    try:
        exams = load_exams()
        exams = [e for e in exams if e.exam_id != exam_id]
        save_exams(exams)
        reset_scheduler()
        return jsonify({'success': True, 'message': 'Exam deleted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/teachers/<teacher_id>', methods=['PUT'])
def api_update_teacher(teacher_id):
    """Update a teacher"""
    try:
        data = request.json
        teachers = load_teachers()
        
        for teacher in teachers:
            if teacher.teacher_id == teacher_id:
                if 'name' in data:
                    teacher.name = data['name']
                if 'title' in data:
                    teacher.title = data['title']
                if 'phone' in data:
                    teacher.phone = data['phone']
                if 'department' in data:
                    teacher.department = data['department']
                break
        
        save_teachers(teachers)
        reset_scheduler()
        return jsonify({'success': True, 'message': 'Teacher updated successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/exams/<exam_id>', methods=['PUT'])
def api_update_exam(exam_id):
    """Update an exam"""
    try:
        data = request.json
        exams = load_exams()
        
        for exam in exams:
            if exam.exam_id == exam_id:
                if 'name' in data:
                    exam.exam_name = data['name']
                if 'subject' in data:
                    exam.subject = data['subject']
                if 'date' in data:
                    exam.date = data['date']
                if 'time_slot' in data:
                    exam.time_slot = data['time_slot']
                if 'room' in data:
                    exam.room = data['room']
                if 'required_teachers' in data:
                    exam.required_teachers = int(data['required_teachers'])
                if 'rooms_count' in data:
                    exam.rooms_count = int(data.get('rooms_count', 6))
                break
        
        save_exams(exams)
        reset_scheduler()
        return jsonify({'success': True, 'message': 'Exam updated successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/teachers/export')
def api_export_teachers():
    """Export teachers to Excel"""
    try:
        teachers = load_teachers()
        data = []
        for t in teachers:
            data.append({
                '工号': t.teacher_id,
                '姓名': t.name,
                '职称': t.title,
                '联系方式': t.phone,
                '所属部门': t.department,
                '监考次数': t.exam_count
            })
        df = pd.DataFrame(data)
        df.to_excel(config.TEACHERS_FILE, index=False)
        return send_file(
            config.TEACHERS_FILE,
            as_attachment=True,
            download_name=f'teachers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/exams/export')
def api_export_exams():
    """Export exams to Excel"""
    try:
        exams = load_exams()
        data = []
        for e in exams:
            data.append({
                '考试编号': e.exam_id,
                '考试名称': e.exam_name,
                '科目': e.subject,
                '日期': e.date,
                '时间段': e.time_slot,
                '考场': e.room,
                '需要监考人数': e.required_teachers
            })
        df = pd.DataFrame(data)
        df.to_excel(config.EXAMS_FILE, index=False)
        return send_file(
            config.EXAMS_FILE,
            as_attachment=True,
            download_name=f'exams_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/teachers/clear', methods=['POST'])
def api_clear_teachers():
    """Clear all teachers"""
    try:
        import config
        df = pd.DataFrame()
        df.to_excel(config.TEACHERS_FILE, index=False)
        reset_scheduler()
        return jsonify({'success': True, 'message': 'All teachers cleared'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/exams/clear', methods=['POST'])
def api_clear_exams():
    """Clear all exams"""
    try:
        import config
        df = pd.DataFrame()
        df.to_excel(config.EXAMS_FILE, index=False)
        reset_scheduler()
        return jsonify({'success': True, 'message': 'All exams cleared'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/schedule', methods=['POST'])
def api_schedule():
    """Execute scheduling"""
    try:
        global scheduler_instance
        teachers = load_teachers()
        exams = load_exams()
        config = load_config()
        
        if not teachers or not exams:
            return jsonify({'success': False, 'error': 'No data available'}), 400
        
        scheduler_instance = ExamScheduler(teachers, exams, config)
        
        schedules = scheduler_instance.schedule()
        
        schedule_list = []
        for s in schedules:
            teachers = [{'id': t.teacher_id, 'name': t.name} for t in s.teachers]
            schedule_list.append({
                'exam_id': s.exam.exam_id,
                'exam_name': s.exam.exam_name,
                'subject': s.exam.subject,
                'date': s.exam.date,
                'time_slot': s.exam.time_slot,
                'room': s.exam.room,
                'teachers': teachers,
                'teacher_count': len(s.teachers)
            })
        
        return jsonify({
            'success': True,
            'data': schedule_list,
            'count': len(schedule_list),
            'message': f'Successfully scheduled {len(schedule_list)} exams'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/schedule')
def api_get_schedule():
    """Get current schedule"""
    try:
        scheduler = get_scheduler()
        if scheduler is None or not scheduler.final_schedules:
            return jsonify({'success': True, 'data': [], 'count': 0, 'message': 'No schedule yet'})

        schedule_list = []
        for s in scheduler.final_schedules:
            teachers = [{'id': t.teacher_id, 'name': t.name} for t in s.teachers]
            schedule_list.append({
                'exam_id': s.exam.exam_id,
                'exam_name': s.exam.exam_name,
                'subject': s.exam.subject,
                'date': s.exam.date,
                'time_slot': s.exam.time_slot,
                'room': s.exam.room,
                'teachers': teachers,
                'teacher_count': len(s.teachers)
            })

        return jsonify({
            'success': True,
            'data': schedule_list,
            'count': len(schedule_list)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/statistics')
def api_statistics():
    """Get statistics"""
    try:
        scheduler = get_scheduler()
        if scheduler is None or not scheduler.final_schedules:
            return jsonify({
                'success': True,
                'total_exams': 0,
                'scheduled_exams': 0,
                'unscheduled_exams': 0,
                'teacher_stats': [],
                'date_stats': {}
            })


        stats = scheduler.get_statistics()
        return jsonify({'success': True, 'stats': stats})
    except Exception as e:
        import traceback
        print(f"统计错误: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/statistics/export')
def api_export_statistics():
    """Export statistics to Excel with charts (as images like page display)"""
    try:
        scheduler = get_scheduler()
        if scheduler is None or not scheduler.final_schedules:
            return jsonify({'success': False, 'error': 'No statistics to export'}), 400

        stats = scheduler.get_statistics()

        import matplotlib.pyplot as plt
        import matplotlib
        matplotlib.use('Agg')

        # 使用绝对路径创建临时图片目录
        img_dir = os.path.join(config.DATA_DIR, 'temp_images')
        if not os.path.exists(img_dir):
            os.makedirs(img_dir)

        # 使用绝对路径保存Excel文件
        filename = os.path.join(config.DATA_DIR, f"statistics_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        teacher_stats = stats['teacher_stats']
        date_stats = stats['date_stats']

        # 保存老师柱状图（类似页面展示）
        if teacher_stats:
            names = [t['name'] for t in teacher_stats]
            counts = [t['exam_count'] for t in teacher_stats]

            plt.figure(figsize=(12, 6))
            bars = plt.bar(names, counts, color='#667eea')
            plt.xlabel('监考老师', fontsize=12, fontweight='bold')
            plt.ylabel('监考次数', fontsize=12, fontweight='bold')
            plt.title('老师监考次数统计', fontsize=14, fontweight='bold', pad=20)
            plt.xticks(rotation=45, ha='right')
            plt.grid(axis='y', alpha=0.3, linestyle='--', color='#ccc')

            for bar in bars:
                height = bar.get_height()
                plt.text(bar.get_x() + bar.get_width()/2., height,
                         f'{int(height)} 场',
                         ha='center', va='bottom', fontsize=10, fontweight='bold')

            plt.tight_layout()
            plt.savefig(os.path.join(img_dir, 'teacher_chart.png'),
                       dpi=150, bbox_inches='tight', facecolor='white', edgecolor='none')
            plt.close()

        # 保存每日排班统计图（类似页面展示）
        if date_stats:
            dates = sorted(date_stats.keys())
            counts = [date_stats[d]['count'] for d in dates]

            plt.figure(figsize=(12, 6))
            for i, (date, count) in enumerate(zip(dates, counts)):
                plt.bar(date, count, color='#667eea')

            plt.xlabel('日期', fontsize=12, fontweight='bold')
            plt.ylabel('排班场次', fontsize=12, fontweight='bold')
            plt.title('每日排班统计', fontsize=14, fontweight='bold', pad=20)
            plt.xticks(rotation=45, ha='right')
            plt.grid(axis='y', alpha=0.3, linestyle='--', color='#ccc')

            for i, (date, count) in enumerate(zip(dates, counts)):
                plt.text(i, count, f'{count} 场', ha='center', va='bottom',
                        fontsize=10, fontweight='bold')

            plt.tight_layout()
            plt.savefig(os.path.join(img_dir, 'date_chart.png'),
                       dpi=150, bbox_inches='tight', facecolor='white', edgecolor='none')
            plt.close()

        # 保存饼图
        scheduled = stats['scheduled_exams']
        unscheduled = stats['unscheduled_exams']
        if scheduled > 0 or unscheduled > 0:
            labels = ['已排班', '未排班']
            sizes = [scheduled, unscheduled]
            colors = ['#667eea', '#dc3545']

            plt.figure(figsize=(8, 8))
            plt.pie(sizes, labels=labels, colors=colors,
                   autopct='%1.1f%%', shadow=True, startangle=90,
                   textprops={'fontsize': 12, 'fontweight': 'bold'})
            plt.title('排班完成情况', fontsize=14, fontweight='bold', pad=20)
            plt.savefig(os.path.join(img_dir, 'pie_chart.png'),
                       dpi=150, bbox_inches='tight', facecolor='white', edgecolor='none')
            plt.close()

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.drawing.image import Image

        wb = Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')

        # Sheet 1: 老师监考统计图
        ws1 = wb.create_sheet('老师监考统计图')
        if teacher_stats and os.path.exists(os.path.join(img_dir, 'teacher_chart.png')):
            img = Image(os.path.join(img_dir, 'teacher_chart.png'))
            img.width = 1200
            img.height = 600
            ws1.add_image(img, 'A1')

        # Sheet 2: 每日排班统计图
        ws2 = wb.create_sheet('每日排班统计图')
        if date_stats and os.path.exists(os.path.join(img_dir, 'date_chart.png')):
            img2 = Image(os.path.join(img_dir, 'date_chart.png'))
            img2.width = 1200
            img2.height = 600
            ws2.add_image(img2, 'A1')

        # Sheet 3: 总体统计（饼图）
        ws3 = wb.create_sheet('总体统计图')
        if stats['scheduled_exams'] > 0 and os.path.exists(os.path.join(img_dir, 'pie_chart.png')):
            img3 = Image(os.path.join(img_dir, 'pie_chart.png'))
            img3.width = 600
            img3.height = 600
            ws3.add_image(img3, 'A1')

        wb.save(filename)

        # 清理临时图片
        try:
            for f in os.listdir(img_dir):
                if f.endswith('.png'):
                    os.remove(os.path.join(img_dir, f))
            os.rmdir(img_dir)
        except:
            pass

        return send_file(
            os.path.abspath(filename),
            as_attachment=True,
            download_name=f'statistics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        import traceback
        print(f"导出统计错误: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export')
def api_export():
    """Export schedule to Excel (default format: horizontal)"""
    try:
        scheduler = get_scheduler()
        if scheduler is None or not scheduler.final_schedules:
            return jsonify({'success': False, 'error': 'No schedule to export'}), 400

        # 默认使用横向考场格式导出
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return export_excel_horizontal(scheduler, timestamp)
    except Exception as e:
        import traceback
        print(f"导出错误: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ============ 多种Excel导出格式 API ============

@app.route('/api/schedule/excel/formats')
def api_schedule_excel_formats():
    """Get available Excel export format options"""
    formats = [
        {
            'id': 'horizontal',
            'name': '横向考场格式（默认）',
            'description': '科目横向排列，每个考场一列，最推荐的格式',
            'preview': '日期 | 时间 | 科目 | 考场一 | 考场二 | ... | 考场六'
        },
        {
            'id': 'standard',
            'name': '标准表格格式',
            'description': '简洁的行式表格，适合数据查看和编辑',
            'preview': '日期 | 时间 | 考场 | 科目 | 监考老师'
        },
        {
            'id': 'printable',
            'name': '横版打印格式',
            'description': '适合A4横版打印，有表头样式',
            'preview': '带表头边框和专业字体的打印样式'
        },
        {
            'id': 'summary',
            'name': '汇总统计格式',
            'description': '排班表 + 老师统计 + 负责人签名区',
            'preview': '排班数据 + 统计信息 + 签名区'
        }
    ]
    return jsonify({'success': True, 'data': formats})


@app.route('/api/schedule/excel/<format>')
def api_schedule_excel(format):
    """Export schedule in specific Excel format"""
    try:
        scheduler = get_scheduler()
        if scheduler is None or not scheduler.final_schedules:
            return jsonify({'success': False, 'error': 'No schedule to export'}), 400

        os.makedirs(config.DATA_DIR, exist_ok=True)
        
        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format == 'horizontal':
            return export_excel_horizontal(scheduler, timestamp)
        else:
            return jsonify({'success': False, 'error': f'Format {format} not implemented yet'}), 500
            
    except Exception as e:
        import traceback
        print(f"Excel导出错误: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500


def export_excel_horizontal(scheduler, timestamp):
    """横向考场格式 - 科目横向排列，每个考场一列"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict
    
    filename = os.path.join(config.DATA_DIR, f'schedule_horizontal_{timestamp}.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = '排班表'
    
    # 样式
    title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_fill = PatternFill(start_color='92C5DE', end_color='92C5DE', fill_type='solid')
    border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), 
                     top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
    
    # 标题
    ws.merge_cells('A1:I1')
    ws['A1'] = '监考排班表'
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 表头
    headers = ['日期', '时间', '科目', '考场一', '考场二', '考场三', '考场四', '考场五', '考场六']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 按日期、时间、科目分组
    grouped = defaultdict(list)
    for schedule in scheduler.final_schedules:
        key = (schedule.exam.date, schedule.exam.time_slot, schedule.exam.subject)
        grouped[key].append(schedule)
    
    # 填充数据
    row = 4
    for (date, time_slot, subject), schedules in sorted(grouped.items()):
        # 填充日期、时间、科目列
        ws.cell(row=row, column=1, value=date).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=2, value=time_slot).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=row, column=3, value=subject).alignment = Alignment(horizontal='center', vertical='center')
        
        # 提取考场并填充
        rooms = {}
        for schedule in schedules:
            room_name = schedule.exam.room
            # 提取考场编号
            room_num = room_name.replace(schedule.exam.subject, '').replace('考场', '')
            if not room_num:
                room_num = '1'
            elif room_num[-1].isdigit():
                room_num = room_num[-1]
            room_key = int(room_num) if room_num.isdigit() else min(6, max(1, len(schedule.exam.room)))
            teacher_names = '、'.join(t.name for t in sorted(schedule.teachers, key=lambda x: x.teacher_id))
            rooms[room_key] = teacher_names
        
        # 填充考场列 (1-6)
        for num in range(1, 7):
            cell = ws.cell(row=row, column=3 + num)
            cell.value = rooms.get(num, '')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        row += 1
    
    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    for i in range(4, 10):
        ws.column_dimensions[get_column_letter(i)].width = 20
    
    # 保存
    wb.save(filename)
    
    return send_file(filename, as_attachment=True,
                    download_name=f'排班表_横向考场_{timestamp}.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/reset', methods=['POST'])
def api_reset():
    """Reset data and schedule"""
    try:
        reset_scheduler()
        return jsonify({'success': True, 'message': 'Reset successful'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


if __name__ == '__main__':
    init_data_dir()

    # 打包后使用debug=False，避免显示调试信息
    debug_mode = not getattr(sys, 'frozen', False)

    # 自动打开浏览器
    import threading
    import webbrowser
    import time

    def open_browser():
        # 等待2秒让服务器启动
        time.sleep(2)
        webbrowser.open('http://localhost:5000')

    # 在后台线程中打开浏览器，不阻塞服务器启动
    threading.Thread(target=open_browser, daemon=True).start()

    app.run(debug=debug_mode, host='0.0.0.0', port=5000)
